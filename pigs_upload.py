# -*- coding: utf-8 -*-

# https://ppg.naif.org.tw/naif/MarketInformation/Pig/TranStatistics.aspx?fbclid=IwAR2lDfuw_FlZBSFmYP52iZ-umNxKEWn66FwRQ8enQFmrdTYNuc-zdfgEZqM 畜產行情資訊網

# 市場: 彰化縣 -> 查詢，撈 --規格豬(75~155公斤) : 頭數、平均價格

# 結果放 excel，每日18點更新

# 引入必要套件
import re
import time
import requests
import openpyxl
from bs4 import BeautifulSoup

class PIGS():

    def __init__(self):
        # Session
        self.session = requests.Session()

        get_date = time.localtime(time.time())
        self.date_list = [get_date.tm_year, get_date.tm_mon, get_date.tm_mday]
        self.date = '-'.join(str(date) for date in self.date_list)

    def get_data(self):
        url = 'https://ppg.naif.org.tw/naif/MarketInformation/Pig/TranStatistics.aspx?fbclid=IwAR2lDfuw_FlZBSFmYP52iZ-umNxKEWn66FwRQ8enQFmrdTYNuc-zdfgEZqM'
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Content-Type': 'application/x-www-form-urlencoded',
        }
        data = {
            '__EVENTTARGET': '',
            '__EVENTARGUMENT': '',
            '__VIEWSTATE': '/wEPDwUKLTU3MTAwMTA5Nw9kFgJmD2QWAmYPZBYCAgMPZBYEAgEPZBYCAgEPZBYEAg0PDxYCHgRUZXh0BRvluILloLTliKXllq7ml6XkuqTmmJPooYzmg4VkZAIPD2QWDGYPZBYKZg8PZBYCHghSZWFkT25seQUJUmVhZHlPbmx5ZAIBDxAPFgYeDURhdGFUZXh0RmllbGQFCXNob3J0TmFtZR4ORGF0YVZhbHVlRmllbGQFBGNvZGUeC18hRGF0YUJvdW5kZ2QQFRYJ5paw5YyX5biCCeWunOiYree4ownmoYPlnJLluIIJ5paw56u557ijCeiLl+agl+e4ownlj7DkuK3luIIJ5aSn5a6J5Y2ACeW9sOWMlue4ownljZfmipXnuKMJ6Zuy5p6X57ijCeWYiee+qeW4ggnlmInnvqnnuKMM6Ie65Y2X5a6J5Y2XCeiHuuWNl+W4ggnpq5jpm4TluIIM6auY6ZuE5bKh5bGxDOmrmOmbhOmzs+WxsQzpq5jpm4Tml5flsbEJ5bGP5p2x57ijCeWPsOadsee4ownoirHok67nuKMJ5r6O5rmW57ijFRYESDIzOARIMjY4BEgzMzgESDMwMgRIMzU2BEg0MDAESDQzOQRINTE0BEg1NDAESDYzMgRINjAwBEg2MTMESDcwMARINzQxBEg4MDAESDgyNQRIODMwBEg4NDIESDkwMARIOTMwBEg5NTUESDg4MBQrAxZnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZGQCAw8PFgIeB1Zpc2libGVnZBYKAgEPDxYCHwAFKjIwMjEtMDgtMzE8YnIvPuOAgOOAgOOAgOOAgOOAgOi+sjExMC0wNy0yNGRkAgMPDxYCHwAFCeW9sOWMlue4o2RkAgUPDxYCHwAFBDE4OTJkZAIHDw8WAh8ABQQxMzkzZGQCCQ8PFgIfAAUEMTkzMGRkAgQPPCsAEQMADxYEHwRnHgtfIUl0ZW1Db3VudAIOZAEQFgAWABYADBQrAAAWAmYPZBYeAgEPZBYOZg8PFgIfAAUOLS3miJDkuqTnuL3mlbhkZAIBDw8WAh8ABQQxODkyZGQCAg8PFgIfAAUCLTRkZAIDDw8WAh8ABQUxMTYuOGRkAgQPDxYCHwAFATRkZAIFDw8WAh8ABQU4MS43NmRkAgYPDxYCHwAFATBkZAICD2QWDmYPDxYCHwAFGS0t6KaP5qC86LGsKDk1fjE1NeWFrOaWpClkZAIBDw8WAh8ABQQxNjM0ZGQCAg8PFgIfAAUCLTFkZAIDDw8WAh8ABQUxMjAuNGRkAgQPDxYCHwAFATNkZAIFDw8WAh8ABQU4My4yMWRkAgYPDxYCHwAFAi0xZGQCAw9kFg5mDw8WAh8ABQs3NX45NeWFrOaWpGRkAgEPDxYCHwAFAzE0NGRkAgIPDxYCHwAFAy0xMWRkAgMPDxYCHwAFBDg4LjhkZAIEDw8WAh8ABQExZGQCBQ8PFgIfAAUFNzIuNTVkZAIGDw8WAh8ABQExZGQCBA9kFg5mDw8WAh8ABQw5NX4xMTXlhazmlqRkZAIBDw8WAh8ABQM1OTRkZAICDw8WAh8ABQMtMjJkZAIDDw8WAh8ABQUxMDcuM2RkAgQPDxYCHwAFATBkZAIFDw8WAh8ABQU4Mi45MGRkAgYPDxYCHwAFATBkZAIFD2QWDmYPDxYCHwAFDTExNX4xMzXlhazmlqRkZAIBDw8WAh8ABQM4MTZkZAICDw8WAh8ABQEzZGQCAw8PFgIfAAUFMTI0LjFkZAIEDw8WAh8ABQExZGQCBQ8PFgIfAAUFODMuNjhkZAIGDw8WAh8ABQItMWRkAgYPZBYOZg8PFgIfAAUNMTM1fjE1NeWFrOaWpGRkAgEPDxYCHwAFAzIyNGRkAgIPDxYCHwAFAzExN2RkAgMPDxYCHwAFBTE0MS43ZGQCBA8PFgIfAAUBMGRkAgUPDxYCHwAFBTgyLjM1ZGQCBg8PFgIfAAUCLTNkZAIHD2QWDmYPDxYCHwAFDzE1NeWFrOaWpOS7peS4imRkAgEPDxYCHwAFAThkZAICDw8WAh8ABQMtMzhkZAIDDw8WAh8ABQUxNjIuNmRkAgQPDxYCHwAFATFkZAIFDw8WAh8ABQU3Ny41OGRkAgYPDxYCHwAFATBkZAIID2QWDmYPDxYCHwAFDjc15YWs5pak5Lul5LiLZGQCAQ8PFgIfAAUCMzBkZAICDw8WAh8ABQMtNDVkZAIDDw8WAh8ABQQ2OC4xZGQCBA8PFgIfAAUBNGRkAgUPDxYCHwAFBTQ3LjI0ZGQCBg8PFgIfAAUBOWRkAgkPZBYOZg8PFgIfAAUM5reY5rGw56iu6LGsZGQCAQ8PFgIfAAUBM2RkAgIPDxYCHwAFAzIwMGRkAgMPDxYCHwAFBTE4My4zZGQCBA8PFgIfAAUCLTdkZAIFDw8WAh8ABQU2MS44NmRkAgYPDxYCHwAFATVkZAIKD2QWDmYPDxYCHwAFCeWFtuS7luixrGRkAgEPDxYCHwAFAjczZGQCAg8PFgIfAAUDLTEwZGQCAw8PFgIfAAUFMTAzLjlkZAIEDw8WAh8ABQE0ZGQCBQ8PFgIfAAUFNzEuMDFkZAIGDw8WAh8ABQIxMmRkAgsPZBYOZg8PFgIfAAUJ5Ya35YeN5bugZGQCAQ8PFgIfAAUDMjc4ZGQCAg8PFgIfAAUCNDNkZAIDDw8WAh8ABQUxMjUuM2RkAgQPDxYCHwAFAThkZAIFDw8WAh8ABQU3OS4zOGRkAgYPDxYCHwAFAi0xZGQCDA9kFg5mDw8WAh8ABR8tLeaIkOS6pOe4veaVuCjkuI3lkKvlhrflh43lu6ApZGQCAQ8PFgIfAAUEMTYxNGRkAgIPDxYCHwAFBiZuYnNwO2RkAgMPDxYCHwAFBTExNS40ZGQCBA8PFgIfAAUGJm5ic3A7ZGQCBQ8PFgIfAAUFODIuMjBkZAIGDw8WAh8ABQYmbmJzcDtkZAIND2QWDmYPDxYCHwAFBuWFrOixrGRkAgEPDxYGHwAFBDEwNDQeCUJhY2tDb2xvcgoAHgRfIVNCAghkZAICDw8WBh8ABQIxNh8HCgAfCAIIZGQCAw8PFgYfAAUFMTE4LjIfBwoAHwgCCGRkAgQPDxYGHwAFATUfBwoAHwgCCGRkAgUPDxYGHwAFBTc5LjcxHwcKAB8IAghkZAIGDw8WBh8ABQEwHwcKAB8IAghkZAIOD2QWDmYPDxYCHwAFBuavjeixrGRkAgEPDxYGHwAFAzg0OB8HCgAfCAIIZGQCAg8PFgYfAAUDLTIwHwcKAB8IAghkZAIDDw8WBh8ABQUxMTUuMh8HCgAfCAIIZGQCBA8PFgYfAAUBMh8HCgAfCAIIZGQCBQ8PFgYfAAUFODQuMzMfBwoAHwgCCGRkAgYPDxYGHwAFATEfBwoAHwgCCGRkAg8PDxYCHwVoZGQCBQ8PFgIfBWdkZAIBD2QWBmYPDxYCHwAFCjIwMjEtMDktMDIWAh8BBQlSZWFkeU9ubHlkAgMPPCsAEQIBEBYAFgAWAAwUKwAAZAIEDzwrABECARAWABYAFgAMFCsAAGQCAg9kFgpmDw8WAh8ABQoyMDIxLTA4LTI3FgIfAQUJUmVhZHlPbmx5ZAIBDw8WAh8ABQoyMDIxLTA5LTAyFgIfAQUJUmVhZHlPbmx5ZAICDxAPFgYfAgUJc2hvcnROYW1lHwMFBGNvZGUfBGdkEBUWCeaWsOWMl+W4ggnlrpzomK3nuKMJ5qGD5ZyS5biCCeaWsOeruee4ownoi5fmoJfnuKMJ5Y+w5Lit5biCCeWkp+WuieWNgAnlvbDljJbnuKMJ5Y2X5oqV57ijCembsuael+e4ownlmInnvqnluIIJ5ZiJ576p57ijDOiHuuWNl+WuieWNlwnoh7rljZfluIIJ6auY6ZuE5biCDOmrmOmbhOWyoeWxsQzpq5jpm4Tps7PlsbEM6auY6ZuE5peX5bGxCeWxj+adsee4ownlj7DmnbHnuKMJ6Iqx6JOu57ijCea+jua5lue4oxUWBEgyMzgESDI2OARIMzM4BEgzMDIESDM1NgRINDAwBEg0MzkESDUxNARINTQwBEg2MzIESDYwMARINjEzBEg3MDAESDc0MQRIODAwBEg4MjUESDgzMARIODQyBEg5MDAESDkzMARIOTU1BEg4ODAUKwMWZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZxYBZmQCBQ88KwARAgEQFgAWABYADBQrAABkAgYPPCsAEQIBEBYAFgAWAAwUKwAAZAIDD2QWBmYPDxYCHwAFCjIwMjEtMDktMDIWAh8BBQlSZWFkeU9ubHlkAgEPEA8WBh8CBQlzaG9ydE5hbWUfAwUEY29kZR8EZ2QQFRYJ5paw5YyX5biCCeWunOiYree4ownmoYPlnJLluIIJ5paw56u557ijCeiLl+agl+e4ownlj7DkuK3luIIJ5aSn5a6J5Y2ACeW9sOWMlue4ownljZfmipXnuKMJ6Zuy5p6X57ijCeWYiee+qeW4ggnlmInnvqnnuKMM6Ie65Y2X5a6J5Y2XCeiHuuWNl+W4ggnpq5jpm4TluIIM6auY6ZuE5bKh5bGxDOmrmOmbhOmzs+WxsQzpq5jpm4Tml5flsbEJ5bGP5p2x57ijCeWPsOadsee4ownoirHok67nuKMJ5r6O5rmW57ijFRYESDIzOARIMjY4BEgzMzgESDMwMgRIMzU2BEg0MDAESDQzOQRINTE0BEg1NDAESDYzMgRINjAwBEg2MTMESDcwMARINzQxBEg4MDAESDgyNQRIODMwBEg4NDIESDkwMARIOTMwBEg5NTUESDg4MBQrAxZnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnFgFmZAIEDzwrABECARAWABYAFgAMFCsAAGQCBA9kFgoCAQ8QZA8WFWYCAQICAgMCBAIFAgYCBwIIAgkCCgILAgwCDQIOAg8CEAIRAhICEwIUFhUQBQMxMTAFBDIwMjFnEAUDMTA5BQQyMDIwZxAFAzEwOAUEMjAxOWcQBQMxMDcFBDIwMThnEAUDMTA2BQQyMDE3ZxAFAzEwNQUEMjAxNmcQBQMxMDQFBDIwMTVnEAUDMTAzBQQyMDE0ZxAFAzEwMgUEMjAxM2cQBQMxMDEFBDIwMTJnEAUDMTAwBQQyMDExZxAFAjk5BQQyMDEwZxAFAjk4BQQyMDA5ZxAFAjk3BQQyMDA4ZxAFAjk2BQQyMDA3ZxAFAjk1BQQyMDA2ZxAFAjk0BQQyMDA1ZxAFAjkzBQQyMDA0ZxAFAjkyBQQyMDAzZxAFAjkxBQQyMDAyZxAFAjkwBQQyMDAxZxYBZmQCAw8QZA8WDGYCAQICAgMCBAIFAgYCBwIIAgkCCgILFgwQBQExBQExZxAFATIFATJnEAUBMwUBM2cQBQE0BQE0ZxAFATUFATVnEAUBNgUBNmcQBQE3BQE3ZxAFATgFAThnEAUBOQUBOWcQBQIxMAUCMTBnEAUCMTEFAjExZxAFAjEyBQIxMmcWAQIIZAIFDxBkZBYBZmQCBw8QDxYGHwIFCXNob3J0TmFtZR8DBQRjb2RlHwRnZBAVFgnmlrDljJfluIIJ5a6c6Jit57ijCeahg+WckuW4ggnmlrDnq7nnuKMJ6IuX5qCX57ijCeWPsOS4reW4ggnlpKflronljYAJ5b2w5YyW57ijCeWNl+aKlee4ownpm7LmnpfnuKMJ5ZiJ576p5biCCeWYiee+qee4owzoh7rljZflronljZcJ6Ie65Y2X5biCCemrmOmbhOW4ggzpq5jpm4TlsqHlsbEM6auY6ZuE6bOz5bGxDOmrmOmbhOaXl+WxsQnlsY/mnbHnuKMJ5Y+w5p2x57ijCeiKseiTrue4ownmvo7muZbnuKMVFgRIMjM4BEgyNjgESDMzOARIMzAyBEgzNTYESDQwMARINDM5BEg1MTQESDU0MARINjMyBEg2MDAESDYxMwRINzAwBEg3NDEESDgwMARIODI1BEg4MzAESDg0MgRIOTAwBEg5MzAESDk1NQRIODgwFCsDFmdnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2cWAWZkAg0PPCsAEQIBEBYAFgAWAAwUKwAAZAIFD2QWBmYPDxYCHwAFCjIwMjEtMDktMDIWAh8BBQlSZWFkeU9ubHlkAgMPPCsAEQIBEBYAFgAWAAwUKwAAZAIEDzwrABECARAWABYAFgAMFCsAAGQCAw8PFgIfAAUJOSwxMzgsMjc1ZGQYCgVPY3RsMDAkY3RsMDAkQ29udGVudFBsYWNlSG9sZGVyX2NvbnRhbnQkQ29udGVudFBsYWNlSG9sZGVyX2NvbnRhbnQkR3JpZFZpZXcyX1BERg9nZAVLY3RsMDAkY3RsMDAkQ29udGVudFBsYWNlSG9sZGVyX2NvbnRhbnQkQ29udGVudFBsYWNlSG9sZGVyX2NvbnRhbnQkR3JpZFZpZXc0D2dkBUtjdGwwMCRjdGwwMCRDb250ZW50UGxhY2VIb2xkZXJfY29udGFudCRDb250ZW50UGxhY2VIb2xkZXJfY29udGFudCRHcmlkVmlldzUPZ2QFT2N0bDAwJGN0bDAwJENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JEdyaWRWaWV3Nl9QREYPZ2QFUGN0bDAwJGN0bDAwJENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JE11bHRpVmlld19tYWluDw9kZmQFS2N0bDAwJGN0bDAwJENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JEdyaWRWaWV3MQ88KwAMAQgCAWQFS2N0bDAwJGN0bDAwJENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JENvbnRlbnRQbGFjZUhvbGRlcl9jb250YW50JEdyaWRWaWV3Ng9nZAVLY3RsMDAkY3RsMDAkQ29udGVudFBsYWNlSG9sZGVyX2NvbnRhbnQkQ29udGVudFBsYWNlSG9sZGVyX2NvbnRhbnQkR3JpZFZpZXczD2dkBU9jdGwwMCRjdGwwMCRDb250ZW50UGxhY2VIb2xkZXJfY29udGFudCRDb250ZW50UGxhY2VIb2xkZXJfY29udGFudCRHcmlkVmlldzNfUERGD2dkBUtjdGwwMCRjdGwwMCRDb250ZW50UGxhY2VIb2xkZXJfY29udGFudCRDb250ZW50UGxhY2VIb2xkZXJfY29udGFudCRHcmlkVmlldzIPZ2SD90Q9TygB4Vnwq/LS0gZ+9jQrV6t1X5aDVSr8NDx7zw==',
            '__VIEWSTATEGENERATOR': '267496D8',
            '__EVENTVALIDATION': '/wEdACDNsImz9qWX0imyTzgUYyH43oIO3TOPL7TCxLMptzdRj12ImlrdCkX9t1/bqhh6qqbKr9WgoTDD+7Zcbib2tSeitlw6BB7DIMWNbkcT1imODDxS9/aNd6jXezAfRo7HSzZwM5CTBOu/EzViT7/ZpTzgZJTyiOAeN1rE0uwTJaBNZBx4xudYxrY4AiAdp7N5c85bz43uMJZp1wscDMTT+DzZOIJcEKMVncuMqyhvhzmcBxxMTpSkyURwXul7iM5HilS4f9Ux3aCOSWDnKb/ZENwHWxDXXhHbn7lb2a92eTlxT3hVrarWVAu1Dr+rcrAMthmhoPuL98ltkoHGGWidm/OmTPsEP0F4tYanf6IJ+uhPqqRPvudQdmAcvRcXBLRWfiR0B5p45w7/LgNh4CxKy7c4pe7STt65ubaDNWB6ye5MUArN04O35Ru8zlEieBpeLcgHjXCzbRUeAWoaHA7n93FSt1uhagzQ3jGHlgRveKj3n0JvxSCNKi38WX2LLhbLqYZCIhOzfynn8o3NvhAm96rOzVS6DTrMTaUV7QVYYyGnUJdlwmeXxvyUgFTLRlPL3ekiV8qMQx2t9QtB/E7pPUAoJInKS/epl1EPWQLJuyQuwgiDDm2hej4hadocUgeoZLC84mGPxzJlPdO5PG8PObOk1xb12qnBPIRPWQ4CVp0NOIBvPOVH08RSvFcolaxlpki83g3B78Us0FozGF3z2+ik',
            'ctl00$ctl00$ContentPlaceHolder_contant$ContentPlaceHolder_contant$TextBox_Content1_QueryDate': self.date,
            'ctl00$ctl00$ContentPlaceHolder_contant$ContentPlaceHolder_contant$DropDownList_Content1_QueryMarket': 'H514',
            'ctl00$ctl00$ContentPlaceHolder_contant$ContentPlaceHolder_contant$Button_Content1_Submit': '查詢'
        }
        response = self.session.post(url, headers=headers, data=data)
        error = re.search('查無資料', response.text)
        correct = re.search('selected="selected" value="H514"', response.text)
        if error:
            amount = '沒資料'
            price = 'https://ppg.naif.org.tw/naif/MarketInformation/Pig/TranStatistics.aspx?fbclid=IwAR2lDfuw_FlZBSFmYP52iZ-umNxKEWn66FwRQ8enQFmrdTYNuc-zdfgEZqM'
        elif correct:
            soup = BeautifulSoup(response.text, 'html.parser')
            tables = soup.find_all("table")[2]
            tds = tables.find_all("td")[8:13]
            amount = tds[0].text
            price = tds[4].text
        else:
            amount = '壞了'
            price = '請洽處理人員'

        return amount, price

    def parse(self):
        date = self.date[5:]
        amount, price = self.get_data()
        workbook_name = str(self.date_list[0]) + '_' + 'pigs.xlsx'
        worksheet_name = str(self.date_list[1]) + '月'
        content_title = ['日期', '頭數', '平均價']
        row1 = 'A' + str(self.date_list[2] +1)
        row2 = 'B' + str(self.date_list[2] +1)
        row3 = 'C' + str(self.date_list[2] +1)
        lists = [str(date), str(amount), str(price)]
        rows = [row1, row2, row3]
        
        try:
            workbook = openpyxl.load_workbook(workbook_name)
            worksheet = workbook.active
        except:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = worksheet_name
            worksheet.append(content_title)

        if worksheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(worksheet_name, 0)
            worksheet.append(content_title)
        
        for row, li in zip(rows, lists):
            worksheet[row] = li
        workbook.save(workbook_name)
        

if __name__ == "__main__":
    pigs = PIGS()
    pigs.parse()
